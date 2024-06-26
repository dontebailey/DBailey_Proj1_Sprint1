import re
from openpyxl import load_workbook
import k_secrets
from serpapi import GoogleSearch
from typing import Tuple, List
from pathlib import Path


def get_data(page: int) -> List[dict]:
    params = {
        "api_key": k_secrets.api_key,
        "engine": "google_jobs",
        "q": "Software Developer",
        "google_domain": "google.com",
        "hl": "en",
        "gl": "us",
        "location": "Boston, Massachusetts, United States",
        "start": page,
        "lrad": "100"
    }

    search = GoogleSearch(params)
    results = search.get_dict()
    return results["jobs_results"]


def get_multiple_pages_of_jobs(num_pages: int) -> list[Tuple]:
    complete_data = []
    for page in range(num_pages):
        current_data = get_data(page)
        clean_data = clean_data_for_db(current_data)
        complete_data.extend(clean_data)
    return complete_data


def clean_data_for_db(raw_job_data: list[dict]) -> list[Tuple]:
    """this is a DRY violation, but I want it to be easy
     to follow, so I'll put it in here for now. There
     should really be one canonical location to the database structure"""
    db_ready_data = []
    for job in raw_job_data:
        job_id = job['job_id']
        job_title = job["title"]
        company_name = job["company_name"]
        job_description = job["description"]
        location = job["location"]
        posted_date = ""
        remote = False
        optional_job_data = job['detected_extensions']
        if optional_job_data.get('posted_at'):
            posted_date = optional_job_data['posted_at']
        if optional_job_data.get('work_from_home'):
            remote = True  # A little optimistic, but all of my data only has work_from_home when TRUE
        url = job['related_links'][0]['link']  # related_lists is a list of dictionaries
        job_highlights = job['job_highlights']
        benefits_section = {}
        # the benefits section can be in and position in the job_highlights list, so we look for it
        for section in job_highlights:
            if section.get("title") == "Benefits":
                benefits_section = section
        min_salary, max_salary = get_salary(benefits_section, job_description.lower())
        salary_time_period = 'N/A'
        if 0 < min_salary < 900:
            salary_time_period = 'Hourly'
        elif min_salary > 0:
            salary_time_period = "Yearly"
        prepared_data = (job_id, job_title, company_name, job_description, location, min_salary, max_salary,
                         salary_time_period, posted_date, url, remote)
        db_ready_data.append(prepared_data)
    return db_ready_data


def get_salary(benefits_section: dict, job_description: str):
    """this is more complicated than you were required to do, I'm looking in several places for salary info"""
    min_salary = 0
    max_salary = 0
    if benefits_section:  # if we got a dictionary with stuff in it
        for benefit_item in benefits_section['items']:
            if 'range' in benefit_item.lower():
                # from https://stackoverflow.com/questions/63714217/how-can-i-extract-numbers-containing-commas-from
                # -strings-in-python
                numbers = re.findall(r'\b\d{1,3}(?:,\d{3})*(?:\.\d+)?(?!\d)', benefit_item)
                if numbers:  # if we found salary data, return it
                    return int(numbers[0].replace(',', '')), int(numbers[1].replace(',', ''))
            numbers = re.findall(r'\b\d{1,3}(?:,\d{3})*(?:\.\d+)?(?!\d)', benefit_item)
            if len(numbers) == 2 and int(
                    numbers[0].replace(',', '')) > 30:  # some jobs just put the numbers in one item
                # and the description in another
                return int(numbers[0].replace(',', '')), int(numbers[1].replace(',', ''))
            else:
                return min_salary, max_salary
    location = job_description.find("salary range")
    if location < 0:
        location = job_description.find("pay range")
    if location < 0:
        return min_salary, max_salary
    numbers = re.findall(r'\b\d{1,3}(?:,\d{3})*(?:\.\d+)?(?!\d)', job_description[location:location + 50])
    if numbers:
        return int(numbers[0].replace(',', '')), int(numbers[1].replace(',', ''))
    return min_salary, max_salary


def get_excel_data() -> list[Tuple]:
    workbook = load_workbook(filename=Path(__file__).parent / Path("Sprint3Data.xlsx"))
    sheet = workbook.active

    max_num_rows = sheet.max_row
    jobs = []
    # sets boundaries for the iteration, so you get one tuple element per row
    for row in sheet.iter_rows(min_row=2,
                               max_row=max_num_rows,
                               min_col=1,
                               max_col=10,
                               values_only=True):
        job_id = row[2]
        job_title = row[9]
        company_name = row[0]
        job_description = "N/A"
        location = row[4]
        min_salary = row[7]
        max_salary = row[6]
        salary_time = row[8]
        posted_at = row[1]
        url = "N/A"
        remote = "N/A"

        prepared_data = (
            job_id, job_title, company_name, job_description, location, min_salary,
            max_salary, salary_time, posted_at, url, remote)

        jobs.append(prepared_data)

    return jobs
